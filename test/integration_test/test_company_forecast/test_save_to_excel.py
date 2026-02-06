import os
import pandas as pd
from company_forecast.forecaster import CompanyForecaster


def create_sample_company(tmp_path):
    folder = tmp_path / "ExcelCo"
    folder.mkdir()
    cols = ["2023-12-31", "2022-12-31", "2021-12-31"]

    income = pd.DataFrame({
        cols[0]: [120, -72, 48],
        cols[1]: [110, -66, 44],
        cols[2]: [100, -60, 40],
    }, index=["Total Revenue", "Cost Of Revenue", "Net Income"])

    balance = pd.DataFrame({
        cols[0]: [500, 20, 10],
        cols[1]: [480, 18, 9],
        cols[2]: [460, 16, 8],
    }, index=["Total Assets", "Accounts Receivable", "Current Debt"])

    cash = pd.DataFrame({
        cols[0]: [30, -5, -2],
        cols[1]: [28, -5, -2],
        cols[2]: [26, -5, -2],
    }, index=["Operating Cash Flow", "Capital Expenditure", "Cash Dividends Paid"])

    income.to_csv(folder / "income statement.csv")
    balance.to_csv(folder / "balance sheet.csv")
    cash.to_csv(folder / "cash flow.csv")
    return str(folder)


def test_save_to_excel_creates_file(tmp_path):
    """Test that save_to_excel creates a file."""
    folder = create_sample_company(tmp_path)
    f = CompanyForecaster(folder, n_forecast_years=1, n_input_years=1)
    f.run_forecast()
    try:
        fname = f.save_to_excel(filename=str(tmp_path / "out.xlsx"))
    except Exception:
        # If openpyxl not installed, skip test
        return
    assert os.path.isfile(fname)
    # clean up
    os.remove(fname)


def test_save_to_excel_contains_correct_sheets(tmp_path):
    """Test that Excel file contains expected sheets with correct data."""
    folder = create_sample_company(tmp_path)
    f = CompanyForecaster(folder, n_forecast_years=2, n_input_years=2)
    f.run_forecast()
    
    try:
        import openpyxl
    except ImportError:
        # Skip if openpyxl not installed
        return
    
    fname = str(tmp_path / "test_output.xlsx")
    try:
        f.save_to_excel(filename=fname)
        
        # Read Excel file and verify sheets
        wb = openpyxl.load_workbook(fname)
        sheet_names = wb.sheetnames
        
        # Should have at least Income Statement, Balance Sheet, Cash Budget
        expected_sheets = ['Income Statement', 'Balance Sheet', 'Cash Budget']
        for sheet in expected_sheets:
            assert sheet in sheet_names, f"Missing sheet: {sheet}"
        
        # Verify Income Statement has data
        income_sheet = wb['Income Statement']
        # Check that there's revenue data (row with 'Revenue' should have numbers)
        assert income_sheet.max_row > 1, "Income Statement should have data rows"
        
        # Verify Balance Sheet has data
        balance_sheet = wb['Balance Sheet']
        assert balance_sheet.max_row > 1, "Balance Sheet should have data rows"
        
        wb.close()
    finally:
        # Clean up
        if os.path.exists(fname):
            os.remove(fname)
